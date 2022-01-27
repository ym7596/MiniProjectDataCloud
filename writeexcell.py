#!/usr/bin/env python
# coding: utf-8

# In[1]:


import os
import openpyxl as op
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import Border, Side
from datetime import datetime
import pandas as pd


def drawBorder(ws, start_c, start_r, end_c, end_r):
    end_c = end_c + 1
    end_r = end_r + 1
    box = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    for i in range(start_c, end_c):
        for j in range(start_r, end_r):
            ws.cell(column = i, row = j).border = box
            
def drawBox(ws, start_c, start_r, end_c, end_r):
    end_r = end_r + 1
    
    box = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    ws.cell(column = start_c, row = start_r).border = Border(top=Side(style='thin'), left=Side(style='thin'))
    ws.cell(column = start_c, row = end_r).border = Border(bottom=Side(style='thin'), left=Side(style='thin'))
    ws.cell(column = end_c, row = start_r).border = Border(top=Side(style='thin'), right=Side(style='thin'))
    ws.cell(column = end_c, row = end_r).border = Border(bottom=Side(style='thin'), right=Side(style='thin'))
    
    for i in range(start_c+1, end_c):
        ws.cell(column = i, row = start_r).border = Border(top=Side(style='thin'))
    for i in range(start_c+1, end_c):
        ws.cell(column = i, row = end_r).border = Border(bottom=Side(style='thin'))
    for i in range(start_r+1, end_r):
        ws.cell(column = start_c, row = i).border = Border(left=Side(style='thin'))
    for i in range(start_r+1, end_r):
        ws.cell(column = end_c, row = i).border = Border(right=Side(style='thin'))


def savepng(kospi, kospi200, kosdaq, name, thema):
    
    lista = thema.T.values.tolist()
    
    themadf = {
        "themaname": lista[0],
        "lastday": lista[1],
        "updown": lista[2]
    }
    
    print(themadf)
    
    pertxt ='''PER은 현재주가에서 주당순이익을 나눈 값으로 낮을수록 저평가된 종목으로 본다.\n일반적으로 10 이하의 PER을 가진 종목을 유망한 종목으로 판단한다.'''
    epstxt = '''주당순이익(EPS)은 기업의 순이익을 유통주식수로 나눈 값을 의미한다.\nEPS에 10을 곱하여 적정주가를 구한 후 현재 주가와 비교하는 방법으로 유망한 종목을 판단할 수 있다.'''
    epsroetxt = '''위에서 언급된 EPS에 자기자본이익율(ROE)를 곱하여 유망 종목을 판단한다.\n해당 값이 높을수록 저평가된 종목으로 판단한다.'''
    pbrtxt = '''주가순자산비율(PBR)은 PER에 ROE를 곱하여 구한다.\nPBR 1이하일 경우 장부가치보다 주가가 낮다고 판단, 따라서 저평가된 종목으로 판단한다.'''
    totaltxt = '''위에서 언급한 네 가지의 방법으로 구해진 각각의 유망한 종목들의 총점을 비교해 유망 종목인 '''+ name + "을 선별"
    predicttxt = name + '''의 지난 주가를 회귀분석을 사용하여 이틀 뒤의 주가를 예측'''
    
    
    wb = op.Workbook()
    ws = wb.active
    ws.sheet_view.showGridLines = False
    
    path = "./imgs"
    img_file = ["per.png", "eps.png", "epsroe.png", "pbr.png", "total.png"]
    txt_list = [pertxt, epstxt, epsroetxt, pbrtxt, totaltxt]
    
    ws.column_dimensions['A'].width = 24.2
    
    ws.merge_cells(start_column = 1, end_column = 7, start_row = 1,end_row = 1)
    ws.cell(column = 1, row = 1).value = "SK 쉴더스 모듈 프로젝트"
    ws.cell(column = 1, row = 1).font = Font(size = 20, bold = True)
    ws.cell(column = 1, row = 1).alignment = Alignment(horizontal = 'center')
    
    ws.merge_cells(start_column = 1, end_column = 7, start_row = 2,end_row = 2)
    ws.cell(column = 1, row = 2).value = "Daily Stock Report"
    ws.cell(column = 1, row = 2).font = Font(size = 15, bold = True)
    ws.cell(column = 1, row = 2).alignment = Alignment(horizontal = 'center')
    
    ws.cell(column = 1, row = 3).value = "Date"
    ws.cell(column = 1, row = 3).font = Font(size = 11, bold = True)
    ws.cell(column = 1, row = 4).value = "Depatment"
    ws.cell(column = 1, row = 4).font = Font(size = 11, bold = True)
    ws.cell(column = 1, row = 5).value = "Participants"
    ws.cell(column = 1, row = 5).font = Font(size = 11, bold = True)
    
    ws.merge_cells(start_column = 2, end_column = 7, start_row = 3,end_row = 3)
    ws.merge_cells(start_column = 2, end_column = 7, start_row = 4,end_row = 4)
    ws.merge_cells(start_column = 2, end_column = 7, start_row = 5,end_row = 5)
    
    ws.cell(column = 2, row = 3).value = datetime.today()
    ws.cell(column = 2, row = 3).alignment = Alignment(horizontal = 'center')
    ws.cell(column = 2, row = 4).value = "5조"
    ws.cell(column = 2, row = 4).alignment = Alignment(horizontal = 'center')
    ws.cell(column = 2, row = 5).value = "박재현, 배지성, 신용민, 최두영, 최하진"
    ws.cell(column = 2, row = 5).alignment = Alignment(horizontal = 'center')
    
    ws.merge_cells(start_column = 1, end_column = 1, start_row = 6,end_row = 8)
    ws.cell(column = 1, row = 6).value = "Stock-Market Indices"
    ws.cell(column = 1, row = 6).font = Font(bold = True)
    ws.cell(column = 1, row = 6).alignment = Alignment(vertical='center')
    
    ws.cell(column = 2, row = 6).value = "KOSPI"
    ws.cell(column = 2, row = 6).font = Font(size = 10)
    ws.cell(column = 3, row = 6).value = kospi
    ws.cell(column = 3, row = 6).font = Font(size = 10)
    ws.cell(column = 3, row = 6).alignment = Alignment(horizontal = 'center')
    
    ws.cell(column = 2, row = 7).value = "KOSPI200"
    ws.cell(column = 2, row = 7).font = Font(size = 10)
    ws.cell(column = 3, row = 7).value = kospi200
    ws.cell(column = 3, row = 7).font = Font(size = 10)
    ws.cell(column = 3, row = 7).alignment = Alignment(horizontal = 'center')
    
    ws.cell(column = 2, row = 8).value = "KOSDAQ"
    ws.cell(column = 2, row = 8).font = Font(size = 10)
    ws.cell(column = 3, row = 8).value = kosdaq
    ws.cell(column = 3, row = 8).font = Font(size = 10)
    ws.cell(column = 3, row = 8).alignment = Alignment(horizontal = 'center')
    
    ws.merge_cells(start_column = 3, end_column = 7, start_row = 6,end_row = 6)
    ws.merge_cells(start_column = 3, end_column = 7, start_row = 7,end_row = 7)
    ws.merge_cells(start_column = 3, end_column = 7, start_row = 8,end_row = 8)
    
    ws.row_dimensions[9].horizontal = 5.1
    ws.merge_cells(start_column = 1, end_column = 7, start_row = 9,end_row = 9)
    ws.cell(column = 1, row = 9).value = "1. 테마"
    ws.cell(column = 1, row = 9).font = Font(size = 13, bold = True)
    
    for i in range(10,16):
        ws.merge_cells(start_column = 2, end_column = 4, start_row = i,end_row = i)
        ws.merge_cells(start_column = 5, end_column = 7, start_row = i,end_row = i)
    
    drawBorder(ws,1,1,7,15)
    
    ws.cell(column = 1, row = 10).value = "테마명"
    ws.cell(column = 2, row = 10).value = "전일대비"
    ws.cell(column = 5, row = 10).value = "최근 3일 등락률(평균)"
    
    ws.cell(column = 1, row = 10).alignment = Alignment(horizontal='center')
    ws.cell(column = 2, row = 10).alignment = Alignment(horizontal='center')
    ws.cell(column = 5, row = 10).alignment = Alignment(horizontal='center')
    
    for i in range(5):
        ws.cell(column = 1, row = 11+i).value = themadf["themaname"][i]
        ws.cell(column = 2, row = 11+i).value = themadf["lastday"][i]
        ws.cell(column = 5, row = 11+i).value = themadf["updown"][i]
        
        ws.cell(column = 1, row = 11+i).alignment = Alignment(horizontal='center')
        ws.cell(column = 2, row = 11+i).alignment = Alignment(horizontal='center')
        ws.cell(column = 5, row = 11+i).alignment = Alignment(horizontal='center')
    
    
    ws.cell(column = 1, row = 16).value = "2. 적정 종목 선택과정"
    ws.cell(column = 1, row = 16).font = Font(size = 13, bold = True)
    
    r = 17
    
    drawBox(ws, 1, 16, 7, r+(13*len(img_file))-1)
    
    for i in range(len(img_file)):
        img = Image(path + "/" + img_file[i])
        ws.add_image(img,"A"+str(r+1))
        
        ws.merge_cells(start_column = 5, start_row = r + 1, end_column = 7, end_row = r + 8)
        ws.cell(column = 5, row = r + 1).alignment = Alignment(wrap_text = True, vertical='top')
        ws.cell(column = 5, row = r + 1).value = txt_list[i]
        
        drawBox(ws,1,r,7,r+12)
        drawBox(ws,5,r,7,r+12)
        
        r = r+13
    
    ws.merge_cells(start_column = 1, end_column = 7, start_row = 83,end_row = 83)
    ws.cell(column = 1, row = 83).value = "3. 유망 종목 예측"
    ws.cell(column = 1, row = 83).font = Font(size = 13, bold = True)
    drawBorder(ws, 1, 83, 7, 83)
    
    img = Image(path + "/" + "img.png")
    ws.add_image(img,"A"+str(84))

    ws.merge_cells(start_column = 5, end_column = 7, start_row = 85,end_row = 91)
    ws.cell(column = 5, row = 85).alignment = Alignment(wrap_text = True, vertical='top')
    ws.cell(column = 5, row = 85).value = predicttxt
    drawBox(ws,1,84,7,95)
    drawBox(ws,5,84,7,95)
 

    wb.save(r"주식 분석 보고서.xlsx")
    wb.close()

